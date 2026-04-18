from datetime import date, datetime, timezone
from io import BytesIO
from zoneinfo import ZoneInfo

import pandas as pd
from openpyxl.styles import Font

_TZ_HU = ZoneInfo("Europe/Budapest")


def _calendar_date_hu(val):
    """Naptári dátum magyar formában; időzóna-észleléssel (UTC → Europe/Budapest)."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    if isinstance(val, pd.Timestamp):
        val = val.to_pydatetime()
    if isinstance(val, datetime):
        dt = val
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(_TZ_HU).strftime("%Y.%m.%d.")
    if isinstance(val, date):
        return val.strftime("%Y.%m.%d.")
    s = str(val).strip()
    if len(s) >= 10 and s[4] == "-" and s[7] == "-":
        if len(s) == 10:
            return s.replace("-", ".") + "."
        try:
            dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
        except ValueError:
            return s[:10].replace("-", ".") + "."
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(_TZ_HU).strftime("%Y.%m.%d.")
    if hasattr(val, "isoformat"):
        return _calendar_date_hu(val.isoformat())
    return s[:10].replace("-", ".") + "." if len(s) >= 10 else s


def format_project_status(status):
    status_map = {"ongoing": "Folyamatban",
                  "completed": "Befejezve", "cancelled": "Megszakítva"}
    return status_map.get(status, status or "")


def get_hours(start_str, end_str, break_min):
    try:
        start = datetime.fromisoformat(str(start_str).replace("Z", "+00:00"))
        end = datetime.fromisoformat(str(end_str).replace("Z", "+00:00"))
        delta = end - start
        return round((delta.total_seconds() / 3600) - (break_min / 60), 2)
    except Exception:
        return 0


def _to_float_or_none(value):
    if value is None or value is False:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        s = value.strip().replace(" ", "")
        if not s:
            return None
        # tizedes: vessző vagy pont (egyszerűsített)
        if "," in s and "." not in s:
            s = s.replace(",", ".")
        try:
            return float(s)
        except ValueError:
            return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def format_project_sheet(ws):
    bold_font = Font(bold=True)
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 20
    for row_idx in [2, 3, 9]:
        for col_idx in range(2, 5):
            ws.cell(row=row_idx, column=col_idx).font = bold_font


def format_materials_sheet(ws, data_len, total_sum):
    bold_font = Font(bold=True)
    red_bold = Font(bold=True, color="FF0000")
    widths = {"B": 15, "C": 35, "D": 12, "E": 10, "F": 15, "G": 15, "H": 20}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for cell in ws[2]:
        if cell.column >= 2:
            cell.font = bold_font
    summary_row = data_len + 3
    ws.cell(row=summary_row, column=2, value="ÖSSZESEN").font = red_bold
    total_cell = ws.cell(row=summary_row, column=8, value=total_sum)
    total_cell.font = red_bold
    total_cell.number_format = "#,##0"


def format_worklog_sheet(ws, data_len, total_hours, total_money):
    bold_font = Font(bold=True)
    red_bold = Font(bold=True, color="FF0000")
    widths = {"B": 12, "C": 20, "D": 15, "E": 15, "F": 25,
              "G": 8, "H": 12, "I": 15}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for cell in ws[2]:
        if cell.column >= 2:
            cell.font = bold_font
    summary_row = data_len + 3
    ws.cell(row=summary_row, column=2, value="MINDÖSSZESEN").font = red_bold
    ws.cell(row=summary_row, column=7, value=total_hours).font = red_bold
    ws.cell(row=summary_row, column=9, value=total_money).font = red_bold


def format_machines_sheet(ws, data_len, machine_totals):
    bold_font = Font(bold=True)
    red_bold = Font(bold=True, color="FF0000")
    widths = {"B": 15, "C": 25, "D": 15, "E": 15, "F": 15}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for cell in ws[2]:
        if cell.column >= 2:
            cell.font = bold_font
    current_row = data_len + 3
    for m_name, total in machine_totals.items():
        ws.cell(row=current_row, column=2, value="ÖSSZESEN").font = red_bold
        ws.cell(row=current_row, column=3, value=m_name).font = red_bold
        ws.cell(row=current_row, column=6, value=total).font = red_bold
        current_row += 1


def build_export_xlsx(data: dict) -> bytes:
    p = data.get("project") or {}
    project_content = [
        ["PROJEKT ADATOK", None],
        ["Projekt Neve", p.get("projectName")],
        ["Megrendelő", p.get("customerName")],
        ["Helyszín", p.get("projectLocation")],
        ["Státusz", format_project_status(p.get("projectStatus"))],
        [None, None],
        ["KAPCSOLAT", "(Ügyfél adatok)"],
        ["Email", p.get("customerEmail")],
        ["Telefon", p.get("customerPhone")],
    ]

    mat_rows, total_mat_sum = [], 0
    for m in data.get("material") or []:
        price_mode = "Egységár" if m.get(
            "priceMode") == "unitPrice" else "Egyedi ár"
        mat_rows.append([
            _calendar_date_hu(m.get("date")),
            m.get("name", ""),
            m.get("quantity", ""),
            m.get("unit", ""),
            m.get("unitPrice", ""),
            price_mode,
            m.get("price", 0),
        ])
        total_mat_sum += float(m.get("price") or 0)

    roles_map = {1: "Admin", 2: "Építésvezető", 3: "Kertész"}
    users_dict = {
        str(u["id"]): u
        for u in (data.get("users") or [])
        if u.get("id") is not None
    }
    wage_type_data = data.get("wageType") or {}
    wage_by_workspace = wage_type_data.get("byWorkspace") or {}
    work_rows, total_work_hours, total_work_money = [], 0, 0

    for log in data.get("worklog") or []:
        emp_raw = log.get("employeeId")
        user = users_dict.get(str(emp_raw), {}) if emp_raw is not None else {}
        hours = get_hours(
            log.get("startTime"),
            log.get("endTime"),
            log.get("breakMinutes") or 0,
        )
        workspace_id = log.get("workspaceId")
        wid_key = (
            str(workspace_id).strip()
            if workspace_id is not None and str(workspace_id).strip() != ""
            else None
        )
        wage_cfg = wage_by_workspace.get(wid_key) or {} if wid_key else {}
        custom_by_uid = wage_cfg.get("customByUid") or {}
        emp_key = str(emp_raw) if emp_raw is not None else None
        salary = (
            _to_float_or_none(custom_by_uid.get(emp_key))
            if emp_key
            else None
        )
        if salary is None:
            salary = _to_float_or_none(wage_cfg.get("defaultValue"))
        if salary is None:
            salary = float(user.get("salary") or 0)
        work_rows.append([
            _calendar_date_hu(log.get("date")),
            user.get("name", "Ismeretlen"),
            roles_map.get(user.get("role"), ""),
            "",
            (log.get("description") or "").replace("\n", " "),
            hours,
            salary,
            hours * salary,
        ])
        total_work_hours += hours
        total_work_money += hours * salary

    machines_lookup = {m["id"]: m.get("name", "Ismeretlen")
                       for m in data.get("machines") or []}
    machine_rows = []
    machine_totals = {}

    for mlog in data.get("machineWorklog") or []:
        m_name = machines_lookup.get(mlog.get("machineId"), "Ismeretlen gép")
        prev = float(mlog.get("previousHours") or 0)
        new_h = float(mlog.get("newHours") or 0)
        daily_usage = new_h - prev
        machine_rows.append([
            _calendar_date_hu(mlog.get("date")),
            m_name,
            prev,
            new_h,
            daily_usage,
        ])
        machine_totals[m_name] = machine_totals.get(m_name, 0) + daily_usage

    df_mat = pd.DataFrame(
        mat_rows,
        columns=["Dátum", "Anyag Megnevezése", "Mennyiség",
                 "Egység", "Egységár (Ft)", "Ár Módja", "VÉGÖSSZEG (Ft)"],
    )
    df_work = pd.DataFrame(
        work_rows,
        columns=["Dátum", "Dolgozó Neve", "Szerepkör", "Munka Típusa", "Leírás",
                 "Óra", "Díj (Ft/óra)", "Összesen (Ft)"],
    )
    df_mach = pd.DataFrame(
        machine_rows,
        columns=["Dátum", "Gép Neve", "Kezdő Óraállás",
                 "Záró Óraállás", "Napi Üzemóra"],
    )

    _date_fmt = "%Y.%m.%d."
    if not df_mat.empty:
        df_mat["Dátum"] = pd.to_datetime(
            df_mat["Dátum"], format=_date_fmt, errors="coerce")
        df_mat = df_mat.sort_values(by="Dátum")
        df_mat["Dátum"] = df_mat["Dátum"].dt.strftime(_date_fmt)
    if not df_work.empty:
        df_work["Dátum"] = pd.to_datetime(
            df_work["Dátum"], format=_date_fmt, errors="coerce")
        df_work = df_work.sort_values(by="Dátum")
        df_work["Dátum"] = df_work["Dátum"].dt.strftime(_date_fmt)
    if not df_mach.empty:
        df_mach["Dátum"] = pd.to_datetime(
            df_mach["Dátum"], format=_date_fmt, errors="coerce")
        df_mach = df_mach.sort_values(by="Dátum")
        df_mach["Dátum"] = df_mach["Dátum"].dt.strftime(_date_fmt)

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        pd.DataFrame(project_content).to_excel(
            writer, sheet_name="Projekt", startrow=1, startcol=1, index=False, header=False
        )
        df_mat.to_excel(writer, sheet_name="Alapanyagok",
                        startrow=1, startcol=1, index=False)
        df_work.to_excel(writer, sheet_name="Munkadíjak",
                         startrow=1, startcol=1, index=False)
        df_mach.to_excel(writer, sheet_name="Munkagépek",
                         startrow=1, startcol=1, index=False)

        format_project_sheet(writer.sheets["Projekt"])
        format_materials_sheet(
            writer.sheets["Alapanyagok"], len(df_mat), total_mat_sum)
        format_worklog_sheet(writer.sheets["Munkadíjak"], len(
            df_work), total_work_hours, total_work_money)
        format_machines_sheet(
            writer.sheets["Munkagépek"], len(df_mach), machine_totals)

    buffer.seek(0)
    return buffer.getvalue()
